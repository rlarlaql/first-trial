# -*- coding: utf-8 -*-
from __future__ import annotations
from fastapi import FastAPI, UploadFile, File, Form, Response
from pydantic import BaseModel
from typing import Dict, List, Optional
from datetime import datetime, timedelta
from collections import defaultdict
import io, re

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ortools.sat.python import cp_model

app = FastAPI(title="Nurse 28-day Cycle Scheduler")

# ------------------------ 유틸 ------------------------
KR_WD = ["일","월","화","수","목","금","토"]
def kr_wd(dt: datetime) -> str:
    return KR_WD[(dt.weekday()+1)%7]
def dlabel(start: datetime, d: int) -> str:
    dt = start + timedelta(days=d-1)
    return f"{dt.month}/{dt.day}({kr_wd(dt)})"
def weeks_28() -> list[list[int]]:
    return [list(range(1,8)), list(range(8,15)), list(range(15,22)), list(range(22,29))]

# ------------------------ 엑셀 저장 ------------------------
def export_excel(schedule_df: pd.DataFrame, start_date: datetime) -> bytes:
    wb = Workbook()
    ws = wb.active; ws.title = "번표"
    ws["A1"] = "이름"
    for d in range(1,29):
        ws[f"{get_column_letter(1+d)}1"] = dlabel(start_date, d)
    ws["AD1"], ws["AE1"], ws["AF1"], ws["AG1"] = "D 간호사 별 갯수","E 간호사 별 갯수","N 간호사 별 갯수","비근무 총합"
    for i,(name,row) in enumerate(schedule_df.iterrows(), start=2):
        ws[f"A{i}"] = name
        for d in range(1,29):
            ws[f"{get_column_letter(1+d)}{i}"] = row[d]
        ws[f"AD{i}"] = f'=COUNTIF(B{i}:AC{i},"D")'
        ws[f"AE{i}"] = f'=COUNTIF(B{i}:AC{i},"E")'
        ws[f"AF{i}"] = f'=COUNTIF(B{i}:AC{i},"N")'
        ws[f"AG{i}"] = (f'=COUNTIF(B{i}:AC{i},"주휴")+COUNTIF(B{i}:AC{i},"*OFF")'
                        f'+COUNTIF(B{i}:AC{i},"생휴")+COUNTIF(B{i}:AC{i},"VAC")+COUNTIF(B{i}:AC{i},"*수면*")')
    # 스타일
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    ws.freeze_panes = "B2"
    for i in range(1,34):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 16 if col in ["AD","AE","AF","AG"] else (12 if col=="A" else 10)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ------------------------ 템플릿 파서 ------------------------
def _to_dt_mmdd(s: str, year_hint: int) -> Optional[datetime]:
    m = re.search(r"(\d{1,2})[^\d]?(\d{1,2})", s)
    if not m: return None
    mm, dd = int(m.group(1)), int(m.group(2))
    return datetime(year_hint, mm, dd)

def parse_template(xlsx_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws_inp = wb["입력"]; ws_prev = wb["이전주기상태"]

    # 주기 시작일(B3): "02월 01일" 또는 "3/1(일)" 등
    start_raw = ws_inp["B3"].value
    if isinstance(start_raw, datetime):
        start_date = start_raw
    elif isinstance(start_raw, str):
        mmdd = _to_dt_mmdd(start_raw, datetime.today().year)
        if not mmdd: raise ValueError("주기 시작일(B3) 형식 인식 실패")
        start_date = mmdd
    else:
        raise ValueError("주기 시작일(B3) 누락")

    # 간호사/주휴 요일
    nurses, weekly_rest = [], {}
    for r in range(13, 25):
        name = ws_inp[f"B{r}"].value
        wd   = ws_inp[f"D{r}"].value
        if name:
            name = str(name).strip(); nurses.append(name)
            weekly_rest[name] = (str(wd).strip() if wd else "월")

    # 제한(H-06)
    no_n, off_on_holiday, tue_no_n = set(), set(), set()
    base = 26
    for r in range(base+2, base+2+20):
        nm = ws_inp[f"A{r}"].value
        typ= ws_inp[f"B{r}"].value
        desc = ws_inp[f"C{r}"].value
        if not nm: continue
        nm=str(nm).strip(); d=str(desc or "").strip()
        if "N 불가" in d: no_n.add(nm)
        if ("법정공휴일" in d) and ("OFF" in d or "오프" in d): off_on_holiday.add(nm)
        if "매주 화" in d and ("D/E만" in d or "D만" in d or "E만" in d): tue_no_n.add(nm)

    # 이전주기상태
    previous_state = {}
    r=2
    while True:
        nm = ws_prev[f"A{r}"].value
        if not nm: break
        last = ws_prev[f"B{r}"].value or "OFF"
        cons = int(ws_prev[f"C{r}"].value or 0)
        ended= (str(ws_prev[f"D{r}"].value).strip()=="예")
        rem  = int(ws_prev[f"E{r}"].value or 0)
        previous_state[str(nm).strip()] = {"last":str(last).strip(), "consN":cons, "ended":ended, "remain_off":rem}
        r+=1

    return {
        "start_date": start_date,
        "nurses": nurses,
        "weekly_rest": weekly_rest,
        "no_n": no_n,
        "off_on_holiday": off_on_holiday,
        "tue_no_n": tue_no_n,
        "previous_state": previous_state
    }

# ------------------------ CP-SAT 스케줄러 ------------------------
def solve_schedule(
    start_date: datetime,
    nurses: List[str],
    weekly_rest: Dict[str,str],
    daily_requirements: Dict[str,int],
    previous_state: Dict[str,dict],
    public_holidays: List[datetime],
    rules: Dict,
    holiday_req_override: Optional[Dict[int,Dict[str,int]]] = None
):
    DAYS=28; SHIFTS=["D","E","N","OFF","주휴"]; NUR=range(len(nurses)); DAY=range(1,DAYS+1)

    def d2date(d): return start_date + timedelta(days=d-1)
    def is_tue(d): return kr_wd(d2date(d))=="화"
    def is_pub(d): return any((h.date()==d2date(d).date()) for h in public_holidays)

    model=cp_model.CpModel()
    x={(n,d,s):model.NewBoolVar(f"x_{n}_{d}_{s}") for n in NUR for d in DAY for s in SHIFTS}

    # 하루 1칸
    for n in NUR:
        for d in DAY:
            model.Add(sum(x[(n,d,s)] for s in SHIFTS)==1)

    # H-01 일별 정원(공휴일 예외 가능)
    for d in DAY:
        rD=daily_requirements["D"]; rE=daily_requirements["E"]; rN=daily_requirements["N"]
        if holiday_req_override and d in holiday_req_override:
            ov=holiday_req_override[d]
            rD, rE, rN = ov.get("D",rD), ov.get("E",rE), ov.get("N",rN)
        model.Add(sum(x[(n,d,"D")] for n in NUR)==rD)
        model.Add(sum(x[(n,d,"E")] for n in NUR)==rE)
        model.Add(sum(x[(n,d,"N")] for n in NUR)==rN)

    # H-07 주휴
    if not rules.get("allow_weekly_rest_shift_within_week", False):
        for n, nurse in enumerate(nurses):
            wd=weekly_rest[nurse]
            for d in DAY:
                if kr_wd(d2date(d))==wd:
                    model.Add(x[(n,d,"주휴")]==1)
    else:
        for n,_ in enumerate(nurses):
            for wk in weeks_28():
                model.Add(sum(x[(n,d,"주휴")] for d in wk)==1)

    # H-06 개인제한
    for n, nurse in enumerate(nurses):
        if nurse in rules.get("no_n", set()):
            for d in DAY: model.Add(x[(n,d,"N")]==0)
        if nurse in rules.get("off_on_public_holiday", set()):
            for d in DAY:
                if is_pub(d):
                    model.Add(x[(n,d,"D")]==0); model.Add(x[(n,d,"E")]==0); model.Add(x[(n,d,"N")]==0)
        if nurse in rules.get("tuesday_no_n", set()):
            for d in DAY:
                if is_tue(d): model.Add(x[(n,d,"N")]==0)

    # H-05 금지전환(N->D, E->D)
    for n in NUR:
        for d in range(1,DAYS):
            model.Add(x[(n,d,"E")]+x[(n,d+1,"D")]<=1)
            model.Add(x[(n,d,"N")]+x[(n,d+1,"D")]<=1)

    # H-04 6일 이상 연속근무 금지
    max_cons=rules.get("max_consecutive_work",5)
    for n in NUR:
        for s in range(1, DAYS-6+2):
            model.Add(sum(x[(n,d,"D")]+x[(n,d,"E")]+x[(n,d,"N")] for d in range(s,s+6))<=max_cons)

    # H-02/H-03 Night 블록 2~3 + 종료 후 2일 비근무
    start2={(n,d):model.NewBoolVar(f"s2_{n}_{d}") for n in NUR for d in range(1,28)}
    start3={(n,d):model.NewBoolVar(f"s3_{n}_{d}") for n in NUR for d in range(1,27)}

    for n in NUR:
        for d in DAY:
            covers=[]
            if d in range(1,28): covers.append(start2[(n,d)])
            if d-1 in range(1,28): covers.append(start2[(n,d-1)])
            if d in range(1,27): covers.append(start3[(n,d)])
            if d-1 in range(1,27): covers.append(start3[(n,d-1)])
            if d-2 in range(1,27): covers.append(start3[(n,d-2)])
            if covers:
                model.Add(x[(n,d,"N")]<=sum(covers))
                model.Add(sum(covers)<=1)
    for n in NUR:
        for d in range(1,28):
            if (n,d) in start2:
                model.Add(x[(n,d,"N")]>=start2[(n,d)])
                model.Add(x[(n,d+1,"N")]>=start2[(n,d)])
            if (n,d) in start3:
                model.Add(x[(n,d,"N")]>=start3[(n,d)])
                model.Add(x[(n,d+1,"N")]>=start3[(n,d)])
                model.Add(x[(n,d+2,"N")]>=start3[(n,d)])
    for n in NUR:
        for d in range(1,28):
            if (n,d) in start2:
                endd=d+1
                for k in [1,2]:
                    if endd+k<=28:
                        model.Add(x[(n,endd+k,"D")]==0)
                        model.Add(x[(n,endd+k,"E")]==0)
                        model.Add(x[(n,endd+k,"N")]==0)
            if (n,d) in start3:
                endd=d+2
                for k in [1,2]:
                    if endd+k<=28:
                        model.Add(x[(n,endd+k,"D")]==0)
                        model.Add(x[(n,endd+k,"E")]==0)
                        model.Add(x[(n,endd+k,"N")]==0)

    # 주기 경계 연속성
    for n, nurse in enumerate(nurses):
        st=previous_state.get(nurse, {"last":"OFF","consN":0,"ended":True,"remain_off":0})
        rem=int(st.get("remain_off",0))
        for k in range(1, rem+1):
            if k<=28:
                model.Add(x[(n,k,"D")]==0); model.Add(x[(n,k,"E")]==0); model.Add(x[(n,k,"N")]==0)
        if st.get("last")=="N" and (not st.get("ended",True)) and int(st.get("consN",0))==1:
            model.Add(x[(n,1,"N")]==1)
            if 2<=28: model.Add(x[(n,2,"N")]==1)
            if 3<=28: model.Add(x[(n,3,"D")]==0); model.Add(x[(n,3,"E")]==0); model.Add(x[(n,3,"N")]==0)
            if 4<=28: model.Add(x[(n,4,"D")]==0); model.Add(x[(n,4,"E")]==0); model.Add(x[(n,4,"N")]==0)

    # ------ 소프트 제약 ------
    penalties=[]; W=rules.get("weights",{"balance_shift":5,"balance_off":4,"flow_penalty":3,"rest_adjacent_bonus":-2})
    total_need={s: daily_requirements[s]*28 for s in ["D","E","N"]}
    per_head={s: total_need[s]/len(nurses) for s in ["D","E","N"]}

    for n in NUR:
        for s in ["D","E","N"]:
            cnt=model.NewIntVar(0,28,f"cnt_{n}_{s}")
            model.Add(cnt==sum(x[(n,d,s)] for d in DAY))
            avg_int=int(round(per_head[s]))
            devp=model.NewIntVar(0,28,f"devp_{n}_{s}")
            devm=model.NewIntVar(0,28,f"devm_{n}_{s}")
            model.Add(cnt-avg_int==devp-devm)
            penalties += [W["balance_shift"]*devp, W["balance_shift"]*devm]

    off_avg=9
    for n in NUR:
        cnt_off=model.NewIntVar(0,28,f"off_{n}")
        model.Add(cnt_off==sum(x[(n,d,"OFF")]+x[(n,d,"주휴")] for d in DAY))
        devp=model.NewIntVar(0,28,f"offp_{n}")
        devm=model.NewIntVar(0,28,f"offm_{n}")
        model.Add(cnt_off-off_avg==devp-devm)
        penalties += [W["balance_off"]*devp, W["balance_off"]*devm]

    for n in NUR:
        for d in range(1,28):
            p_ne=model.NewBoolVar(f"pNE_{n}_{d}")
            p_ed=model.NewBoolVar(f"pED_{n}_{d}")
            model.Add(x[(n,d,"N")]+x[(n,d+1,"E")]<=2*p_ne)
            model.Add(x[(n,d,"E")]+x[(n,d+1,"D")]<=2*p_ed)
            penalties += [W["flow_penalty"]*p_ne, W["flow_penalty"]*p_ed]

    for n in NUR:
        for d in DAY:
            if d-1>=1:
                b1=model.NewBoolVar(f"bprev_{n}_{d}")
                model.Add(x[(n,d,"주휴")]+x[(n,d-1,"OFF")]<=2*b1)
                penalties.append(W["rest_adjacent_bonus"]*b1)
            if d+1<=28:
                b2=model.NewBoolVar(f"bnext_{n}_{d}")
                model.Add(x[(n,d,"주휴")]+x[(n,d+1,"OFF")]<=2*b2)
                penalties.append(W["rest_adjacent_bonus"]*b2)

    model.Minimize(sum(penalties))
    solver=cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(rules.get("max_time", 30))
    status=solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None, {"status":"INFEASIBLE", "reason":"하드 제약 충돌(공휴일·주휴·주기경계·N블록)"}

    # 결과 DF
    data=[]
    for n, nurse in enumerate(nurses):
        row={}
        for d in DAY:
            for s in SHIFTS:
                if solver.Value(x[(n,d,s)])==1:
                    row[d]=s; break
        data.append(pd.Series(row, name=nurse))
    df=pd.DataFrame(data)

    # H-01 검증
    shortage={}
    for d in DAY:
        dcnt=(df[d]=="D").sum(); ecnt=(df[d]=="E").sum(); ncnt=(df[d]=="N").sum()
        rD=daily_requirements["D"]; rE=daily_requirements["E"]; rN=daily_requirements["N"]
        if holiday_req_override and d in holiday_req_override:
            o=holiday_req_override[d]; rD,oE,oN = o.get("D",rD), o.get("E",rE), o.get("N",rN); rE=oE; rN=oN
        if not (dcnt==rD and ecnt==rE and ncnt==rN):
            shortage[d]={"D":dcnt,"E":ecnt,"N":ncnt,"req":{"D":rD,"E":rE,"N":rN}}
    return df, {"status":"OK","shortage":shortage}

# ------------------------ 입력 스키마(JSON용) ------------------------
class PrevState(BaseModel):
    last: str = "OFF"
    consN: int = 0
    ended: bool = True
    remain_off: int = 0

class SolvePayload(BaseModel):
    start_date: str
    nurses: List[str]
    weekly_rest: Dict[str,str]
    daily_requirements: Dict[str,int] = {"D":2,"E":2,"N":2}
    previous_state: Dict[str,PrevState]
    public_holidays: List[str] = []
    rules: Dict = {}
    holiday_req_override: Optional[Dict[int,Dict[str,int]]] = None

# ------------------------ 엔드포인트 ------------------------
@app.post("/api/schedule")
async def schedule_api(
    file: UploadFile | None = File(default=None),
    allow_weekly_rest_shift_within_week: Optional[bool] = Form(default=None),
    holiday_override_json: Optional[str] = Form(default=None),
    start_date_json: Optional[str] = Form(default=None),
    json_payload: Optional[str] = Form(default=None)
):
    """
    두 가지 모드 지원:
    1) multipart/form-data + file(.xlsx): 템플릿 파일 업로드 → 자동 배정 → .xlsx 반환
       - allow_weekly_rest_shift_within_week (True/False)
       - holiday_override_json (예: {"16":{"D":1,"E":1,"N":2}, "17":{...}})
    2) application/json: SolvePayload 스키마로 직접 데이터 전송
    """
    if file is not None:
        xlsx_bytes = await file.read()
        parsed = parse_template(xlsx_bytes)
        start_date = parsed["start_date"]
        nurses = parsed["nurses"]
        weekly_rest = parsed["weekly_rest"]
        previous_state = parsed["previous_state"]

        # 기본 정원
        daily_requirements = {"D":2,"E":2,"N":2}

        # 2026-02 시작이면 설 연휴 자동 포함(2/16~18)
        default_holidays = []
        if start_date.year==2026 and start_date.month==2:
            default_holidays = [datetime(2026,2,16), datetime(2026,2,17), datetime(2026,2,18)]

        # 옵션 파싱
        rules = {
            "forbidden_transitions": [("N","D"),("E","D")],
            "max_consecutive_work": 5,
            "no_n": set(parsed["no_n"]),
            "off_on_public_holiday": set(parsed["off_on_holiday"]),
            "tuesday_no_n": set(parsed["tue_no_n"]),
            "allow_weekly_rest_shift_within_week": bool(allow_weekly_rest_shift_within_week) if allow_weekly_rest_shift_within_week is not None else False,
            "weights": {"balance_shift":5,"balance_off":4,"flow_penalty":3,"rest_adjacent_bonus":-2},
            "max_time": 30
        }
        holiday_req_override = None
        if holiday_override_json:
            holiday_req_override = pd.read_json(io.StringIO(holiday_override_json), typ="series").to_dict()

        df, info = solve_schedule(
            start_date=start_date, nurses=nurses, weekly_rest=weekly_rest,
            daily_requirements=daily_requirements, previous_state=previous_state,
            public_holidays=default_holidays, rules=rules,
            holiday_req_override=holiday_req_override
        )
        if df is None:
            return Response(content=str(info).encode("utf-8"), media_type="text/plain", status_code=422)

        xlsx = export_excel(df, start_date)
        fname = f'번표_28일주기_{start_date.strftime("%Y%m%d")}.xlsx'
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'}
        )

    # --- JSON 모드 ---
    # (프론트에서 application/json 로 송신)
    return Response(content=b"Use JSON mode by POSTing to /api/schedule with application/json payload, or send multipart with file=.", media_type="text/plain")

# JSON 모드 전용 핸들러(옵션)
@app.post("/api/schedule-json")
async def schedule_json(payload: SolvePayload):
    start_date = datetime.fromisoformat(payload.start_date)
    public_holidays = [datetime.fromisoformat(x) for x in payload.public_holidays]
    rules = payload.rules or {}
    rules.setdefault("forbidden_transitions", [("N","D"),("E","D")])
    rules.setdefault("max_consecutive_work", 5)
    rules.setdefault("allow_weekly_rest_shift_within_week", False)
    rules.setdefault("weights", {"balance_shift":5,"balance_off":4,"flow_penalty":3,"rest_adjacent_bonus":-2})
    rules.setdefault("max_time", 30)

    prev = {k: v.model_dump() if isinstance(v, PrevState) else v for k,v in payload.previous_state.items()}
    df, info = solve_schedule(
        start_date=start_date, nurses=payload.nurses, weekly_rest=payload.weekly_rest,
        daily_requirements=payload.daily_requirements, previous_state=prev,
        public_holidays=public_holidays, rules=rules,
        holiday_req_override=payload.holiday_req_override
    )
    if df is None:
        return {"status":"INFEASIBLE", "info":info}
    # 엑셀 바이트 → base64 로 돌려도 되지만, 여기선 표만 반환
    return {"status":"OK", "shortage":info.get("shortage", {}), "preview": df.to_dict(orient="index")}
